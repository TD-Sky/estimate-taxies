import openpyxl


class DataSheet():
    """建模数据表格, 有创建和读写功能"""

    def __init__(self, filename):
        self.rows = 200
        self.columns = 15
        self.filename = filename
        self.wb = self._init_wb()
        self.sheet = self.wb['Sheet']


    def _init_wb(self):
        """初始化工作簿"""
        try: 
            wb = openpyxl.load_workbook(self.filename)
        except FileNotFoundError:
            wb = openpyxl.Workbook()

        return wb


    def entitle(self):
        """定制表格标题"""
        self.sheet.merge_cells('A1:A2')
        self.sheet['A1'] = '序号'
        for row in range(3, 203):
            self.sheet[f'A{row}'] = row - 2

        self.sheet.merge_cells('B1:K2')
        self.sheet['B1'] = '样本'

        self.sheet.merge_cells('L1:P1')
        self.sheet['L1'] = '预测'

        for combine in zip('LMNOP', '一二三四五'):
            self.sheet[combine[0]+'2'] = '模型' + combine[1]

        self.wb.save(self.filename)


    def format(self):
        """格式化相应位置"""
        begin = 1
        max_cols = self.rows + 1
        max_rows = self.columns + 2
        thin_border = openpyxl.styles.Border(
                      left = openpyxl.styles.Side(style = 'thin'),
                      right = openpyxl.styles.Side(style = 'thin'),
                      top = openpyxl.styles.Side(style = 'thin'),
                      bottom = openpyxl.styles.Side(style = 'thin')
                      )

        for col in range(begin, max_cols + begin):
            for row in range(begin, max_rows + begin):
                horz = 'left' if col == 1 and row > 2 else 'center'
                cell = self.sheet.cell(row, col)
                cell.alignment = openpyxl.styles.Alignment(horizontal = horz, vertical='center')
                cell.border = thin_border

        self.wb.save(self.filename)



    def write(self, row, sample, sp_size, estimate):
        """把样本与估测结果写入指定行"""
        begin = 2
        end = self.columns + begin
        for col in range(begin, end):
            locat = f'{openpyxl.utils.get_column_letter(col)}{row}'

            if begin <= col <= (sp_size - 1 + begin):
                self.sheet[locat] = sample[col - begin]
            elif (sp_size + begin) <= col <= end:
                self.sheet[locat] = estimate[col - begin - sp_size]
            else:
                raise IndexError(f'No Such i = {i}')

        self.wb.save(self.filename)

